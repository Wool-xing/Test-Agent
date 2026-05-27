# SPDX-License-Identifier: MIT
"""
测试套件减重 v2 — 修复返回类型 + CSV/JSON 输入 + CLI 补全。

改进 (vs suite_minimizer.py):
- minimize_by_coverage 返回类型修正 (docstring → 实现一致)
- 支持 CSV/JSON/Excel 三种输入格式
- CLI 支持 minimize + analyze 子命令
"""

from __future__ import annotations

import csv
import json
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# Text utilities
# ═══════════════════════════════════════════════════════════════

def normalize_text(text: str) -> str:
    t = re.sub(r"[^\w\s一-鿿]", " ", text or "")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def jaccard(a: str, b: str) -> float:
    sa = set(normalize_text(a).split())
    sb = set(normalize_text(b).split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


# ═══════════════════════════════════════════════════════════════
# Duplicate detection
# ═══════════════════════════════════════════════════════════════

def find_duplicates(cases: list[dict],
                    fields: tuple[str, ...] = ("name", "steps", "expected"),
                    threshold: float = 0.85) -> list[dict]:
    def case_text(c: dict) -> str:
        parts = []
        for f in fields:
            v = c.get(f, "")
            if isinstance(v, list):
                v = " ".join(str(x) for x in v)
            parts.append(str(v))
        return " ".join(parts)

    duplicates = []
    n = len(cases)
    for i in range(n):
        for j in range(i + 1, n):
            sim = jaccard(case_text(cases[i]), case_text(cases[j]))
            if sim >= threshold:
                duplicates.append({
                    "case_a": cases[i].get("id", f"#{i}"),
                    "case_b": cases[j].get("id", f"#{j}"),
                    "similarity": round(sim, 3),
                    "name_a": str(cases[i].get("name", "")),
                    "name_b": str(cases[j].get("name", "")),
                })
    return duplicates


# ═══════════════════════════════════════════════════════════════
# Coverage minimization
# ═══════════════════════════════════════════════════════════════

def minimize_by_coverage(test_coverage_map: dict[str, set[tuple[str, int]]]) -> dict[str, Any]:
    """
    Greedy set-cover: minimize test suite while preserving coverage.

    Returns dict with:
      - kept_tests: list of test IDs to keep
      - redundant_tests: list of test IDs that can be removed
      - kept_count: number of kept tests
      - redundant_count: number of redundant tests
      - coverage_lines: total unique (file, line) pairs covered
    """
    all_lines: set = set()
    for cov in test_coverage_map.values():
        all_lines |= cov

    selected: list[str] = []
    covered: set = set()
    remaining = dict(test_coverage_map)

    while covered != all_lines and remaining:
        best_test = max(remaining, key=lambda t: len(remaining[t] - covered))
        new_lines = remaining[best_test] - covered
        if not new_lines:
            break
        selected.append(best_test)
        covered |= new_lines
        del remaining[best_test]

    redundant = [t for t in test_coverage_map if t not in selected]
    return {
        "kept_tests": selected,
        "redundant_tests": redundant,
        "kept_count": len(selected),
        "redundant_count": len(redundant),
        "coverage_lines": len(covered),
    }


# ═══════════════════════════════════════════════════════════════
# Multi-format input
# ═══════════════════════════════════════════════════════════════

def load_cases_csv(path: str) -> list[dict]:
    """Load test cases from CSV (columns: id, name, steps, expected)."""
    cases = []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cases.append({
                "id": row.get("id") or row.get("用例ID", ""),
                "name": row.get("name") or row.get("用例名称", ""),
                "steps": row.get("steps") or row.get("测试步骤", ""),
                "expected": row.get("expected") or row.get("预期结果", ""),
            })
    return cases


def load_cases_json(path: str) -> list[dict]:
    """Load test cases from JSON array."""
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "cases" in data:
        return data["cases"]
    return [data]


def load_cases_excel(path: str) -> list[dict]:
    """Load test cases from Excel (4-sheet format)."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case = dict(zip(headers, row))
        if case.get("用例ID") or case.get("id"):
            cases.append({
                "id": case.get("用例ID") or case.get("id", ""),
                "name": case.get("用例名称") or case.get("name", ""),
                "steps": case.get("测试步骤") or case.get("steps", ""),
                "expected": case.get("预期结果") or case.get("expected", ""),
            })
    return cases


def load_cases(path: str) -> list[dict]:
    """Auto-detect format and load test cases."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return load_cases_excel(path)
    if ext == ".csv":
        return load_cases_csv(path)
    if ext == ".json":
        return load_cases_json(path)
    raise ValueError(f"unsupported format: {ext}")


# ═══════════════════════════════════════════════════════════════
# Combined scan
# ═══════════════════════════════════════════════════════════════

def analyze_suite(path: str, threshold: float = 0.85) -> dict[str, Any]:
    """Load cases from path, find duplicates, return analysis."""
    cases = load_cases(path)
    dups = find_duplicates(cases, threshold=threshold)
    return {
        "source": path,
        "total_cases": len(cases),
        "duplicates_count": len(dups),
        "redundancy_pct": round(len(dups) / max(len(cases), 1) * 100, 1),
        "duplicates": dups[:50],
    }


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO)
    ap = argparse.ArgumentParser(description="Suite minimizer v2")
    sub = ap.add_subparsers(dest="cmd")

    ana = sub.add_parser("analyze", help="Analyze suite for duplicates")
    ana.add_argument("input", help="Test case file (.xlsx/.csv/.json)")
    ana.add_argument("--threshold", type=float, default=0.85)

    mini = sub.add_parser("minimize", help="Minimize based on coverage (requires coverage map JSON)")
    mini.add_argument("--coverage-map", required=True, help="JSON: {test_id: [[file,line],...]}")

    args = ap.parse_args()

    if args.cmd == "analyze":
        result = analyze_suite(args.input, args.threshold)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    elif args.cmd == "minimize":
        raw = json.loads(Path(args.coverage_map).read_text())
        cov_map = {tid: {tuple(x) for x in lines} for tid, lines in raw.items()}
        result = minimize_by_coverage(cov_map)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        ap.print_help()
