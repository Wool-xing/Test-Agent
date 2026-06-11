# SPDX-License-Identifier: MIT
# DEPRECATED: use suite_minimizer_v2 instead. This file will be removed in V1.2.
"""
测试套件减重（Suite Minimization）- 检测重复用例 / 冗余覆盖
被引用方：testcase-designer / 测试质量

策略：
1. 静态：用例文本相似度（标题/步骤/预期）
2. 动态：覆盖率追踪（pytest-cov），保留覆盖独特代码路径的用例
"""
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

logger = logging.getLogger(__name__)


# ===== 静态：用例文本相似度 =====

def normalize_text(text: str) -> str:
    """归一化文本：去标点 / 小写 / 去多空格"""
    t = re.sub(r"[^\w\s一-鿿]", " ", text or "")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def jaccard(a: str, b: str) -> float:
    """字符级 Jaccard 相似度（简化）"""
    sa = set(normalize_text(a).split())
    sb = set(normalize_text(b).split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def find_duplicate_cases(cases: List[Dict],
                          fields: Tuple[str, ...] = ("name", "steps", "expected"),
                          threshold: float = 0.85) -> List[Dict]:
    """
    检测疑似重复用例：合并指定字段计算相似度。
    """
    def case_text(c):
        parts = []
        for f in fields:
            v = c.get(f, "")
            if isinstance(v, list):
                v = " ".join(str(x) for x in v)
            parts.append(str(v))
        return " ".join(parts)

    duplicates = []
    n = len(cases)
    for i in range(n):
        for j in range(i + 1, n):
            sim = jaccard(case_text(cases[i]), case_text(cases[j]))
            if sim >= threshold:
                duplicates.append({
                    "case_a": cases[i].get("id", f"#{i}"),
                    "case_b": cases[j].get("id", f"#{j}"),
                    "similarity": round(sim, 3),
                    "name_a": cases[i].get("name", ""),
                    "name_b": cases[j].get("name", ""),
                })
    return duplicates


# ===== 动态：基于覆盖率减重 =====

def parse_coverage_xml(coverage_xml: str) -> Dict[str, Set[int]]:
    """解析 cobertura coverage.xml，返回 {file: covered_lines}"""
    import xml.etree.ElementTree as ET
    root = ET.parse(coverage_xml).getroot()
    result: Dict[str, Set[int]] = {}
    for cls in root.findall(".//class"):
        filename = cls.get("filename", "")
        for line in cls.findall(".//line"):
            num = int(line.get("number", 0))
            hits = int(line.get("hits", 0))
            if hits > 0:
                result.setdefault(filename, set()).add(num)
    return result


def minimize_by_coverage(test_coverage_map: Dict[str, Set[Tuple[str, int]]]) -> Dict[str, Any]:
    """
    贪心算法：选最少的测试用例覆盖最多的代码行。
    test_coverage_map: {test_id: {(file, line), ...}}
    返回保留的测试 ID 列表（其余可删）。
    """
    all_lines: Set = set()
    for cov in test_coverage_map.values():
        all_lines |= cov

    selected: List[str] = []
    covered: Set = set()
    remaining_tests = dict(test_coverage_map)

    while covered != all_lines and remaining_tests:
        # 选 covers 最多新行的 test
        best_test = max(remaining_tests, key=lambda t: len(remaining_tests[t] - covered))
        new_lines = remaining_tests[best_test] - covered
        if not new_lines:
            break
        selected.append(best_test)
        covered |= new_lines
        del remaining_tests[best_test]

    redundant = [t for t in test_coverage_map if t not in selected]
    return {
        "kept_tests": selected,
        "redundant_tests": redundant,
        "kept_count": len(selected),
        "redundant_count": len(redundant),
        "coverage_lines": len(covered),
    }


# ===== 综合扫描 =====

def scan_excel_for_duplicates(excel_path: str, threshold: float = 0.85) -> Dict:
    """从 4 Sheet Excel 用例文件中检测重复"""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case = dict(zip(headers, row))
        if case.get("用例ID"):
            cases.append({
                "id": case.get("用例ID"),
                "name": case.get("用例名称"),
                "steps": case.get("测试步骤"),
                "expected": case.get("预期结果"),
            })
    dups = find_duplicate_cases(cases, threshold=threshold)
    return {
        "total_cases": len(cases),
        "duplicates_count": len(dups),
        "duplicates": dups[:50],
    }


if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO)
    parser = argparse.ArgumentParser(description="测试套件减重")
    sub = parser.add_subparsers(dest="cmd")
    d = sub.add_parser("dup-excel"); d.add_argument("excel"); d.add_argument("--threshold", type=float, default=0.85)
    args = parser.parse_args()
    if args.cmd == "dup-excel":
        print(json.dumps(scan_excel_for_duplicates(args.excel, args.threshold),
                          indent=2, ensure_ascii=False))
