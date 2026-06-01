# SPDX-License-Identifier: MIT
"""
双向追溯矩阵：需求 ID ↔ 测试用例 ID ↔ 缺陷 ID

被引用方：test-lead / test-coordinator skill / report-generator
输入：PRD 文本 + 测试用例目录 + Bug 草稿 JSON
输出：完整追溯矩阵 (JSON / markdown table)
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# ── ID 提取正则 ──────────────────────────────────────────────
RE_REQUIREMENT = re.compile(r"(?:REQ|FR|NFR|BR)[-_\s]?(\d{2,})", re.IGNORECASE)
RE_TESTCASE = re.compile(r"TC[-_\s]?(?:[A-Z]+[-_\s]?)?(\d{2,})", re.IGNORECASE)
RE_BUG = re.compile(r"(?:BUG|DEFECT|ISSUE)[-_\s]?(\d{2,})", re.IGNORECASE)
RE_XRAY = re.compile(r"XRAY[-_\s]?(\d+)", re.IGNORECASE)


class TraceabilityMatrix:
    """构建双向追溯矩阵。

    用法:
        matrix = TraceabilityMatrix()
        matrix.load_prd("workspace/需求分析/prd.md")
        matrix.load_testcases("workspace/测试用例/")
        matrix.load_bugs("workspace/测试报告/bug_drafts.json")
        report = matrix.build()
    """

    def __init__(self) -> None:
        self._reqs: Dict[str, Dict] = {}       # req_id → {title, text_snippet, linked_tc, linked_bugs}
        self._testcases: Dict[str, Dict] = {}   # tc_id → {title, file, linked_req}
        self._bugs: Dict[str, Dict] = {}        # bug_id → {title, linked_tc, severity}
        self._forward: Dict[str, List[str]] = {}  # req → [tc_ids]
        self._backward: Dict[str, List[str]] = {} # tc → [req_ids]

    # ── 加载 ──────────────────────────────────────────────────

    def load_prd(self, source: str | Path) -> List[str]:
        """从 PRD 文件提取需求 ID，返回 ID 列表。"""
        src = Path(source)
        if not src.exists():
            logger.warning(f"PRD not found: {src}")
            return []
        text = src.read_text(encoding="utf-8", errors="replace")
        ids_found: List[str] = []
        for m in RE_REQUIREMENT.finditer(text):
            rid = m.group(0).upper().replace(" ", "-").replace("_", "-")
            if rid not in self._reqs:
                # 提取标题（ID 所在行）
                line_start = text.rfind("\n", 0, m.start()) + 1
                line_end = text.find("\n", m.end())
                title = text[line_start:line_end if line_end != -1 else len(text)].strip()
                self._reqs[rid] = {
                    "title": title[:120],
                    "text_snippet": text[max(0, m.start() - 40):m.end() + 80].strip(),
                    "linked_tc": [],
                    "linked_bugs": [],
                }
                ids_found.append(rid)
        logger.info(f"PRD {src.name}: {len(ids_found)} requirements found")
        return ids_found

    def load_testcases(self, directory: str | Path) -> int:
        """扫描测试用例目录，通过命名规范 TC-{MODULE}-{TYPE}-{NUM} 提取 ID。"""
        d = Path(directory)
        if not d.is_dir():
            logger.warning(f"testcase dir not found: {d}")
            return 0
        count = 0
        for f in d.rglob("*"):
            if not f.is_file():
                continue
            if f.suffix not in (".py", ".md", ".xlsx", ".json", ".yaml", ".yml"):
                continue
            text = ""
            try:
                if f.suffix in (".py", ".md", ".json", ".yaml", ".yml"):
                    text = f.read_text(encoding="utf-8", errors="replace")
                elif f.suffix == ".xlsx":
                    text = _read_xlsx_text(f)
            except Exception:
                continue

            for m in RE_TESTCASE.finditer(text):
                tc_id = f"TC-{m.group(1).upper()}"
                if tc_id not in self._testcases:
                    self._testcases[tc_id] = {
                        "title": f.name,
                        "file": str(f),
                        "linked_req": [],
                    }
                    count += 1

            # 也检查文件名
            for m in RE_TESTCASE.finditer(f.stem):
                tc_id = f"TC-{m.group(1).upper()}"
                if tc_id not in self._testcases:
                    self._testcases[tc_id] = {
                        "title": f.stem,
                        "file": str(f),
                        "linked_req": [],
                    }
                    count += 1
        logger.info(f"testcases: {count} found in {d}")
        return count

    def load_bugs(self, source: str | Path) -> int:
        """加载 bug 草稿 JSON，提取 BUG ID。"""
        src = Path(source)
        if not src.exists():
            logger.warning(f"bug file not found: {src}")
            return 0
        data = json.loads(src.read_text(encoding="utf-8", errors="replace"))
        bugs_list = data if isinstance(data, list) else data.get("bugs", [data])
        count = 0
        for bug in bugs_list:
            if not isinstance(bug, dict):
                continue
            bug_id = bug.get("id") or bug.get("bug_id") or f"BUG-{hash(str(bug)) % 10000:04d}"
            self._bugs[str(bug_id)] = {
                "title": bug.get("title", bug.get("summary", ""))[:120],
                "linked_tc": bug.get("testcase_id", bug.get("tc_id", "")),
                "severity": bug.get("severity", bug.get("priority", "medium")),
            }
            count += 1
        logger.info(f"bugs: {count} loaded from {src}")
        return count

    # ── 链接 ──────────────────────────────────────────────────

    def link(self) -> Dict:
        """自动链接：需求↔用例↔缺陷。基于 ID 命名规范交叉匹配。"""
        # 需求 ↔ 用例: 模块名匹配
        for rid, rinfo in self._reqs.items():
            req_text = (rinfo.get("title", "") + " " + rinfo.get("text_snippet", "")).lower()
            for tc_id, tcinfo in self._testcases.items():
                tc_file = tcinfo["file"].lower()
                # 用例文件路径中含需求 ID 的数字部分
                req_num = rid.replace("REQ-", "").replace("FR-", "").replace("NFR-", "").replace("BR-", "")
                if req_num in tc_file or req_num in tc_id:
                    if tc_id not in rinfo["linked_tc"]:
                        rinfo["linked_tc"].append(tc_id)
                    if rid not in tcinfo["linked_req"]:
                        tcinfo["linked_req"].append(rid)

        # 缺陷 ↔ 用例: 直接字段引用
        for bug_id, binfo in self._bugs.items():
            linked_tc = binfo.get("linked_tc", "")
            if linked_tc and linked_tc in self._testcases:
                rid_list = self._testcases[linked_tc].get("linked_req", [])
                for rid in rid_list:
                    if bug_id not in self._reqs.get(rid, {}).get("linked_bugs", []):
                        self._reqs.setdefault(rid, {"title": "", "text_snippet": "", "linked_tc": [], "linked_bugs": []})
                        if bug_id not in self._reqs[rid]["linked_bugs"]:
                            self._reqs[rid]["linked_bugs"].append(bug_id)

        return self.build()

    # ── 构建矩阵 ──────────────────────────────────────────────

    def build(self) -> Dict:
        """返回完整双向追溯矩阵。"""
        coverage_total = len(self._reqs)
        coverage_covered = sum(1 for r in self._reqs.values() if r["linked_tc"])
        return {
            "summary": {
                "requirements": len(self._reqs),
                "testcases": len(self._testcases),
                "bugs": len(self._bugs),
                "coverage_pct": round(coverage_covered / max(coverage_total, 1) * 100, 1),
                "uncovered_reqs": [r for r, v in self._reqs.items() if not v["linked_tc"]],
            },
            "forward": {rid: {"title": r["title"], "testcases": r["linked_tc"], "bugs": r["linked_bugs"]}
                        for rid, r in self._reqs.items()},
            "backward": {tc_id: {"title": t["title"], "file": t["file"], "requirements": t["linked_req"]}
                         for tc_id, t in self._testcases.items()},
            "orphan_bugs": [bid for bid, b in self._bugs.items()
                           if not b.get("linked_tc") or b["linked_tc"] not in self._testcases],
        }

    # ── 导出 ──────────────────────────────────────────────────

    def to_markdown(self, report: Optional[Dict] = None) -> str:
        """生成 markdown 追溯矩阵表格。"""
        if report is None:
            report = self.build()
        lines = [
            "# Test Traceability Matrix",
            "",
            f"**Requirements**: {report['summary']['requirements']} | "
            f"**Test Cases**: {report['summary']['testcases']} | "
            f"**Bugs**: {report['summary']['bugs']} | "
            f"**Coverage**: {report['summary']['coverage_pct']}%",
            "",
            "## Forward Trace (Req → TC → Bug)",
            "",
            "| Requirement | Test Cases | Bugs |",
            "|-------------|-----------|------|",
        ]
        for rid, r in report["forward"].items():
            tcs = ", ".join(r["testcases"]) or "—"
            bugs = ", ".join(r["bugs"]) or "—"
            lines.append(f"| {rid} | {tcs} | {bugs} |")
        lines.extend([
            "",
            "## Backward Trace (TC → Req)",
            "",
            "| Test Case | Requirements |",
            "|-----------|-------------|",
        ])
        for tc_id, t in report["backward"].items():
            reqs = ", ".join(t["requirements"]) or "—"
            lines.append(f"| {tc_id} | {reqs} |")
        if report["summary"]["uncovered_reqs"]:
            lines.extend([
                "",
                "## Uncovered Requirements",
                "",
                ", ".join(report["summary"]["uncovered_reqs"]),
            ])
        if report["orphan_bugs"]:
            lines.extend([
                "",
                "## Orphan Bugs (no linked test case)",
                "",
                ", ".join(report["orphan_bugs"]),
            ])
        return "\n".join(lines) + "\n"


# ── helpers ──────────────────────────────────────────────────

def _read_xlsx_text(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        parts = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                parts.append(" ".join("" if v is None else str(v) for v in row))
        return " ".join(parts)
    except Exception:
        return ""


# ── CLI ──────────────────────────────────────────────────────

def main() -> None:
    import argparse
    logging.basicConfig(level=logging.INFO)
    p = argparse.ArgumentParser(description="双向追溯矩阵构建")
    p.add_argument("--prd", help="PRD 文件路径")
    p.add_argument("--testcases", help="测试用例目录")
    p.add_argument("--bugs", help="Bug 草稿 JSON")
    p.add_argument("--link", action="store_true", help="自动链接后输出矩阵")
    p.add_argument("--markdown", action="store_true", help="输出 markdown 表格")
    p.add_argument("--output", help="保存 JSON 到文件")
    args = p.parse_args()

    matrix = TraceabilityMatrix()
    if args.prd:
        matrix.load_prd(args.prd)
    if args.testcases:
        matrix.load_testcases(args.testcases)
    if args.bugs:
        matrix.load_bugs(args.bugs)

    if args.link:
        report = matrix.link()
    else:
        report = matrix.build()

    if args.markdown:
        print(matrix.to_markdown(report))
    else:
        print(json.dumps(report, indent=2, ensure_ascii=False))

    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        Path(args.output).write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        logger.info(f"矩阵已保存: {args.output}")


if __name__ == "__main__":
    main()
