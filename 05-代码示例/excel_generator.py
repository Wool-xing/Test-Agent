# SPDX-License-Identifier: MIT
"""
测试 Excel 生成器（用例/结果）
权威实现：本文件，03-用例设计 agent 文档不再内嵌副本。
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ===== 颜色 =====
COLORS = {
    "header_bg": "366092",
    "header_fg": "FFFFFF",
    "P0": "FF4444",
    "P1": "FF8C00",
    "P2": "4169E1",
    "P3": "808080",
    "pass": "00CC44",
    "fail": "FF2222",
    "skip": "AAAAAA",
    "warning": "FFCC00",
    "row_even": "F5F5F5",
    "row_odd": "FFFFFF",
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_cell(cell, value: str):
    cell.value = value
    cell.font = Font(bold=True, color=COLORS["header_fg"], size=11)
    cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _data_cell(cell, value, row_idx: int = 0, color: Optional[str] = None):
    cell.value = value
    bg = COLORS["row_even"] if row_idx % 2 == 0 else COLORS["row_odd"]
    cell.fill = PatternFill("solid", fgColor=color or bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _priority_cell(cell, priority: str):
    cell.value = priority
    cell.fill = PatternFill("solid", fgColor=COLORS.get(priority, "808080"))
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


# ===== 用例 Excel =====

CASE_HEADERS = [
    ("用例ID", 18), ("模块", 12), ("类型", 8), ("优先级", 10),
    ("用例名称", 28), ("前置条件", 20), ("测试步骤", 38),
    ("测试数据", 22), ("预期结果", 28),
    # API 字段（API 用例时填，UI 用例可留空）
    ("Method", 8), ("Path", 22), ("Headers", 18), ("ExpectedStatus", 12),
    ("实际结果", 20), ("执行状态", 10), ("备注", 18),
]


def _write_case_row(ws, row_idx: int, tc: Dict):
    _data_cell(ws.cell(row=row_idx, column=1), tc.get("id", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=2), tc.get("module", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=3), tc.get("type", "UI"), row_idx)
    _priority_cell(ws.cell(row=row_idx, column=4), tc.get("priority", "P2"))
    _data_cell(ws.cell(row=row_idx, column=5), tc.get("name", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=6), tc.get("precondition", ""), row_idx)
    steps = tc.get("steps", [])
    steps_text = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))
    _data_cell(ws.cell(row=row_idx, column=7), steps_text, row_idx)
    _data_cell(ws.cell(row=row_idx, column=8), tc.get("data", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=9), tc.get("expected", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=10), tc.get("method", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=11), tc.get("path", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=12), tc.get("headers", ""), row_idx)
    _data_cell(ws.cell(row=row_idx, column=13), str(tc.get("expected_status", "")), row_idx)
    _data_cell(ws.cell(row=row_idx, column=14), "", row_idx)
    _data_cell(ws.cell(row=row_idx, column=15), "待执行", row_idx)
    _data_cell(ws.cell(row=row_idx, column=16), tc.get("remark", ""), row_idx)
    ws.row_dimensions[row_idx].height = 40


def _write_summary_sheet(ws, testcases: List[Dict]):
    """Sheet：用例总览（优先级分布 + 模块分布）"""
    p_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    module_counts: Dict[str, int] = {}
    for tc in testcases:
        p = tc.get("priority", "P2")
        p_counts[p] = p_counts.get(p, 0) + 1
        m = tc.get("module", "未分类")
        module_counts[m] = module_counts.get(m, 0) + 1

    ws.cell(row=1, column=1, value="优先级分布").font = Font(bold=True, size=13)
    ws.merge_cells("A1:B1")
    for i, (priority, count) in enumerate(p_counts.items(), 2):
        ws.cell(row=i, column=1, value=priority)
        ws.cell(row=i, column=2, value=count)
    ws.cell(row=7, column=1, value=f"合计：{len(testcases)} 条")

    ws.cell(row=9, column=1, value="模块分布").font = Font(bold=True, size=13)
    ws.merge_cells("A9:B9")
    for i, (m, c) in enumerate(sorted(module_counts.items()), 10):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=c)


def _write_filtered_sheet(ws, cases: List[Dict], title: str):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CASE_HEADERS))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    for col, (h, w) in enumerate(CASE_HEADERS, 1):
        _header_cell(ws.cell(row=2, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    for ri, tc in enumerate(cases, 3):
        _write_case_row(ws, ri, tc)
    ws.freeze_panes = "A3"


def create_testcase_excel(testcases: List[Dict], output_path: str) -> str:
    """
    生成 4 Sheet 测试用例 Excel：
      Sheet1 用例总览（统计）
      Sheet2 测试用例（详细）
      Sheet3 P0冒烟集
      Sheet4 P0+P1回归集
    """
    wb = openpyxl.Workbook()

    # Sheet1: 用例总览
    ws_summary = wb.active
    ws_summary.title = "用例总览"
    _write_summary_sheet(ws_summary, testcases)

    # Sheet2: 详细
    ws_detail = wb.create_sheet("测试用例")
    for col, (h, w) in enumerate(CASE_HEADERS, 1):
        _header_cell(ws_detail.cell(row=1, column=col), h)
        ws_detail.column_dimensions[get_column_letter(col)].width = w
    ws_detail.row_dimensions[1].height = 25
    for ri, tc in enumerate(testcases, 2):
        _write_case_row(ws_detail, ri, tc)
    ws_detail.freeze_panes = "A2"

    # Sheet3: P0冒烟集
    p0 = [tc for tc in testcases if tc.get("priority") == "P0"]
    _write_filtered_sheet(wb.create_sheet("P0冒烟集"), p0, f"P0 冒烟集（共 {len(p0)} 条）")

    # Sheet4: P0+P1回归集
    p01 = [tc for tc in testcases if tc.get("priority") in ("P0", "P1")]
    _write_filtered_sheet(wb.create_sheet("P0_P1回归集"), p01, f"P0+P1 回归集（共 {len(p01)} 条）")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"用例 Excel 已生成: {output_path}（共 {len(testcases)} 条，4 Sheet）")
    return output_path


# ===== 结果 Excel =====

def create_result_excel(results: Dict, output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "执行结果"

    summary = results.get("summary", {})
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="测试执行结果报告")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    summary_data = [
        ("执行时间", summary.get("date", datetime.now().strftime("%Y-%m-%d %H:%M"))),
        ("测试环境", summary.get("environment", "")),
        ("总用例数", summary.get("total", 0)),
        ("通过", summary.get("passed", 0)),
        ("失败", summary.get("failed", 0)),
        ("通过率", f"{summary.get('pass_rate', 0):.1%}"),
    ]
    for ri, (label, value) in enumerate(summary_data, 3):
        ws.cell(row=ri, column=1, value=label).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=value)

    if "failures" in results:
        ws.cell(row=12, column=1, value="失败用例明细").font = Font(bold=True, size=13)
        fail_headers = ["用例ID", "用例名称", "失败类型", "错误信息", "优先级"]
        for col, h in enumerate(fail_headers, 1):
            _header_cell(ws.cell(row=13, column=col), h)
        for ri, failure in enumerate(results["failures"], 14):
            ws.cell(row=ri, column=1, value=failure.get("case_id"))
            ws.cell(row=ri, column=2, value=failure.get("case_name"))
            ws.cell(row=ri, column=3, value=failure.get("failure_type"))
            ws.cell(row=ri, column=4, value=failure.get("error_message", ""))
            _priority_cell(ws.cell(row=ri, column=5), failure.get("priority", "P2"))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ===== 示例 =====

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    sample_cases = [
        {
            "id": "TC-LOGIN-UI-001", "module": "登录", "type": "UI", "priority": "P0",
            "name": "正确账号密码登录",
            "precondition": "用户已注册，账号状态正常",
            "steps": ["打开登录页", "输入账号", "输入密码", "点击登录"],
            "data": "账号:testuser, 密码:Test@123456",
            "expected": "跳转首页，显示用户头像",
        },
        {
            "id": "TC-LOGIN-API-001", "module": "登录", "type": "API", "priority": "P0",
            "name": "登录接口返回 token",
            "method": "POST", "path": "/api/v1/auth/login",
            "headers": "Content-Type: application/json",
            "data": '{"username":"testuser","password":"Test@123456"}',
            "expected_status": 200,
            "expected": "返回 200 + token 字段非空",
        },
        {
            "id": "TC-LOGIN-UI-002", "module": "登录", "type": "UI", "priority": "P1",
            "name": "密码连续错误5次锁定",
            "steps": ["输入正确账号", "输错密码 5 次"],
            "expected": "账号锁定，提示锁定时间",
        },
    ]
    create_testcase_excel(sample_cases, "workspace/测试用例/testcases_sample.xlsx")
